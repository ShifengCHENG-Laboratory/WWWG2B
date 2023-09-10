#2021-6-18 16:36:54
#fengcong@caas.cn
#plot tree in excel,newick tree ,could have bootstrap


# -c --color
# sample1    990099
# sample2    009900
import re
import sys,os
import argparse
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill
# from Bio import Phylo

sys.setrecursionlimit(100000)
def find_comma_index(tree_seq):
    comma_index_list = []
    for index,i in enumerate(tree_seq):
        if i == ",":
            comma_index_list.append(index)

    i_list = []
    for i in comma_index_list:
        if tree_seq[0:i].count("(") == tree_seq[0:i].count(")")  and tree_seq[i+1:].count(")") == tree_seq[i+1:].count("(") :
            # print(111)
            i_list.append(i)

    return i_list
        



def rm_bracket(tree_seq):
    if tree_seq[0] == "(":
        rbound = tree_seq.rfind(")")
        return tree_seq[1:rbound]
    else:
        return tree_seq

def post_traverse_tree_and_plot(tree_seq,depth,plot_sample_list,working_sheet):
    comma_index = find_comma_index(tree_seq)
    comma_index_len = len(comma_index)
    if not comma_index_len:
        pass
    else:
        # left_seq = tree_seq[0:comma_index]
        # left_seq = rm_bracket(left_seq)
        # print(left_seq)
        # # left_seq = ":".join(left_seq.split(":")[0:-1]) if num_ends(left_seq) else left_seq
        # left_leaf_list,l_max_depth,l_row_num = post_traverse_tree_and_plot(left_seq,depth+1,plot_sample_list,working_sheet)
        # # rbound = tree_seq[:-1].rfind(")")
        # right_seq = tree_seq[comma_index+1:]
        # right_seq = rm_bracket(right_seq)
        # print(right_seq)
        # # right_seq = ":".join(right_seq.split(":")[0:-1]) if num_ends(right_seq) else right_seq
        # right_leaf_list,r_max_depth ,r_row_num= post_traverse_tree_and_plot(right_seq,depth+1,left_leaf_list,working_sheet)
        seq_list = []
        pre_start = 0
        for i,indx in enumerate(comma_index) :
            seq_list.append( rm_bracket( tree_seq[pre_start:indx] ) )
            pre_start = indx+1
            if i == len(comma_index) - 1:
                seq_list.append(rm_bracket( tree_seq[pre_start:] ) )
        
        leaf_list_list = []
        max_depth_list = []
        row_num_list = []
        # print(len(seq_list))
        for seq in seq_list:
            plist=[]
            if len(leaf_list_list) ==0:
                plist = plot_sample_list
            else:
                plist = leaf_list_list[-1]
            # print(seq)
            leaf_list,max_depth,row_num= post_traverse_tree_and_plot(seq,depth+1,plist,working_sheet)
            # print(leaf_list)
            leaf_list_list.append(leaf_list)
            max_depth_list.append(max_depth)
            row_num_list.append(row_num)


    if not comma_index_len: # leaf node
        sample_name = tree_seq.split(":")[0]

        # print(sample_name)

        plot_sample_list.append(sample_name)
        row_num = 2*len(plot_sample_list) -1
        col_num = depth
        # border_bot = Border(bottom=Side(border_style="medium",color="000000"))
        # working_sheet.cell(row_num,col_num).border = border_bot
        return (plot_sample_list,depth,row_num)
    else: # parant node
        # row_num_list = len(plot_sample_list) -1
        total_row = 0
        for i in row_num_list:
            total_row += i
        row_num = int( (total_row)/2 )
        col_num = depth
        max_depth=0
        for i in max_depth_list:
            if i > max_depth:
                max_depth = i
        return (leaf_list_list[-1], max_depth, row_num)


def post_traverse_tree_and_plot2(tree_seq,depth,plot_sample_list,working_sheet,max):
    comma_index = find_comma_index(tree_seq)
    comma_index_len = len(comma_index)
    if not comma_index_len:
        #do something
        pass
    else:
        # left_seq = tree_seq[0:comma_index]
        # left_seq = rm_bracket(left_seq)
        # # left_seq = ":".join(left_seq.split(":")[0:-1]) if num_ends(left_seq) else left_seq
        # left_leaf_list,l_col_num,l_row_num = post_traverse_tree_and_plot2(left_seq,depth+1,plot_sample_list,working_sheet,max)
        # right_seq = tree_seq[comma_index+1:]
        # right_seq = rm_bracket(right_seq)
        # # right_seq = ":".join(right_seq.split(":")[0:-1]) if num_ends(right_seq) else right_seq
        # right_leaf_list,r_col_num ,r_row_num= post_traverse_tree_and_plot2(right_seq,depth+1,left_leaf_list,working_sheet,max)
        seq_list = []
        pre_start = 0
        for i,indx in enumerate(comma_index) :
            seq_list.append( rm_bracket( tree_seq[pre_start:indx] ) )
            pre_start = indx+1
            if i == len(comma_index) - 1:
                seq_list.append(rm_bracket( tree_seq[pre_start:] ) )
        
        leaf_list_list = []
        row_num_list = []
        col_num_list = []
        for seq in seq_list:
            plist=[]
            if len(leaf_list_list) ==0:
                plist = plot_sample_list
            else:
                plist = leaf_list_list[-1]
            leaf_list,col_num,row_num= post_traverse_tree_and_plot2(seq,depth+1,plist,working_sheet,max)
            leaf_list_list.append(leaf_list)
            col_num_list.append(col_num)
            row_num_list.append(row_num)

    if not comma_index_len: # leaf node
        sample_name = tree_seq.split(":")[0]
        plot_sample_list.append(sample_name)
        row_num = 2*len(plot_sample_list) -1
        col_num = max
        border_bot = Border(bottom=Side(border_style="medium",color="000000"))
        working_sheet.cell(row_num,col_num).border = border_bot
        return (plot_sample_list,col_num,row_num)
    else: # parant node
        # row_num_list = len(plot_sample_list) -1
        row_num = int( (row_num_list[0]+row_num_list[-1])/2 )
        col_num = min(col_num_list)-1
        

        border_bot = Border(bottom=Side(border_style="medium",color="000000"))
        # print(col_num,depth)
        
        border_rig = Border(right=Side(border_style="medium",color="000000"))
        border_rig_bot = Border(right=Side(border_style="medium",color="000000"),bottom=Side(border_style="medium",color="000000"))
        border_lef_bot = Border(left=Side(border_style="medium",color="000000"),bottom=Side(border_style="medium",color="000000"))
        border_lef_bot_rig = Border(left=Side(border_style="medium",color="000000"),bottom=Side(border_style="medium",color="000000"), \
            right=Side(border_style="medium",color="000000"))

        for row in range(row_num_list[0]+1,row_num_list[-1]+1):
            if row == row_num and col_num != 1:
                working_sheet.cell(row,col_num).border = border_rig_bot
            else:
                working_sheet.cell(row,col_num).border = border_rig
        
        for col in range(col_num+1,col_num_list[0]):
            working_sheet.cell(row_num_list[0],col).border = border_bot
        for i in range(1,len(row_num_list)):
            if col_num +1 == col_num_list[i] and col_num_list[i] != max:
                working_sheet.cell(row_num_list[i],col_num+1).border = border_lef_bot_rig
            else:
                working_sheet.cell(row_num_list[i],col_num+1).border = border_lef_bot
                for col in range(col_num+2,col_num_list[i]):
                    working_sheet.cell(row_num_list[i],col).border = border_bot

        return (leaf_list_list[-1],col_num,row_num)
    


if __name__ == "__main__":
    cmdparser = argparse.ArgumentParser(description="plot tree in excel table/fengcong@caas.cn" )
    cmdparser.add_argument("-o","--output", dest="output",type=str, required=True,
                            help="output xlsx file name")
    cmdparser.add_argument("-c","--color", dest="color",type=str, required=False,
                            help="individual color,sample number/name must constant with the tree file.[optional]")
    cmdparser.add_argument("-i","--insert", dest="insert",type=int, default=2, required=False,
                            help="insert N rows at header.(default=2)[optional]")
    # cmdparser.add_argument("-i","--info", dest="information",type=str, required=False,
    #                         help="individual information,sample number/name must constant with the tree file.[optional]")
    cmdparser.add_argument('trees', metavar='trees', type=str, nargs='+',
                    help='tree files,newick format')

    args = cmdparser.parse_args()

    #deal args
    tree_files_list = args.trees
    output_file_name = args.output
    color_file = args.color
    # tree_files_list = ["/vol2/agis/chengshifeng_group/fengcong/kmer_read_self_test/cYLeLh34Fj8fqALXY43Aaw_newick.txt"]
    # output_file_name= "/vol2/agis/chengshifeng_group/fengcong/kmer_read_self_test/169pure_add_223modern_missense.newick.xlsx"
    # color_file = "/vol2/agis/chengshifeng_group/fengcong/kmer_read_self_test/1059.color"
    # color_file=None

    #open workbook
    tree_workbook = openpyxl.Workbook()

    #recording sheet position
    sheet_pos = 0

    #calc sample count and get sample color 
    
    color_d = {}
    if color_file:
        inf = open(color_file,"r")
        for line in inf.readlines():
            ls = line.strip().split()
            color_d[ls[0]] = ls[1]
        inf.close()


    #traverse each tree file to different Sheet
    for tree_file in tree_files_list: 
        ##open the tree file
        tree_file_handle = open(tree_file,"r")
        lines = tree_file_handle.readlines()
        tree_file_handle.close()
        if len(lines) != 1:
            sys.stderr.write("this file has more than one line.\n")
            exit(-1)
        tree_seq = lines[0].strip()
        tree_seq = tree_seq.strip(";")
        tree_seq = tree_seq[1:-1]  #skip "(" and ")"
        # pattern = re.compile(r'\:\d+\.\d+')
        # tree_seq = re.sub(pattern,"",tree_seq)
        # pattern = re.compile(r'\)\d+\.\d+')
        # tree_seq = re.sub(pattern,")",tree_seq)

        # print(tree_seq.count(")"))
        # print(tree_seq.count("("))
        ## i cant use this package to traverse this tree
        # tree = Phylo.read(tree_file, "newick")

        
        sheet_name = os.path.basename(tree_file).split(".")[0]
        working_sheet = tree_workbook.create_sheet(sheet_name,sheet_pos)

        #we need posttraverse the tree
        sample_list,max_d ,row_num= post_traverse_tree_and_plot(tree_seq,1,[],working_sheet)
        post_traverse_tree_and_plot2(tree_seq,1,[],working_sheet,max_d)
        sys.stdout.write("%d\n"%(max_d)) # if u want to add some information f0llowing this tree, the start col of the information should be maxd+2
        working_sheet.insert_rows(1,args.insert) #insert 2 rows
        algn_center = Alignment(horizontal="center",vertical="center",wrap_text=False)
        light_ft = Font(name='Times New Roman', size=12 , bold=False)
        
        # print(sample_list)
        for index,sample in enumerate(sample_list):
            merge_range_item = [2*(index+1)-1 + args.insert,max_d+1,2*(index+1)+args.insert,max_d+1] #[0,0,0,0] #start_row=None, start_column=None, end_row=None, end_column=None
            working_sheet.merge_cells(start_row=merge_range_item[0], start_column=merge_range_item[1], 
                    end_row=merge_range_item[2], end_column=merge_range_item[3])
            
            sys.stdout.write(sample+"\n")
            if len(color_d) != 0:
                color_fill = PatternFill(fgColor=color_d[sample], fill_type="solid")
                # print(sample,color_d[sample])
                working_sheet.cell(merge_range_item[0],merge_range_item[1]).fill = color_fill
            working_sheet.cell(merge_range_item[0],merge_range_item[1]).value = sample
            working_sheet.cell(merge_range_item[0],merge_range_item[1]).alignment  = algn_center
            working_sheet.cell(merge_range_item[0],merge_range_item[1]).font = light_ft
        sheet_pos +=1

        
    
    tree_workbook.save(filename=output_file_name)
