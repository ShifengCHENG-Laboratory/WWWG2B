#coding:utf-8
#►
#◄

# plot SV and RDV to excel 
# this is the first version 

import sys,os
import gzip
import argparse
from types import WrapperDescriptorType
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill
from openpyxl.utils import cell, get_column_letter
import logging

import re
import subprocess
import copy

logging.basicConfig(level = logging.DEBUG,format = '[%(asctime)s] - %(name)s - %(levelname)s - %(message)s',stream = sys.stderr)
logger = logging.getLogger("SV and RDV 2 Excel")

rdv_data_dir = "/vol3/agis/chengshifeng_group/fengcong/zz.wheat_data.zz/02.1059_depth_data"

Rscript="/public/home/fengcong/anaconda2/envs/R/bin/Rscript"
py3="/public/home/fengcong/anaconda2/envs/py3/bin/python"
bcftools="/public/agis/chengshifeng_group/fengcong/WGRS/software/bcftools1.9/bin/bcftools"
bedtools="/public/agis/chengshifeng_group/fengcong/IPK_assembly/software/bedtools2/bin/bedtools"
part_len="/public/agis/chengshifeng_group/fengcong/WGRS/software/Fc-code/partlen.txt"

def deal_sv(chro,start,end,sv_file,part_len):
    '''{
            "type":"DEL", #DUP INV
            "position":6500,
            "length":700,
            "stat_in_each_sample":[
                "0",
                "0",
                "0",
                "0",
                "0",
                "1"
            ]
        }
    '''
    ##calc the gene+upstream region
    need_region=[start, end]
    

    #chr pos convert to part pos
    need_chr=chro
    # part_len_d={}
    # inf = open(part_len,"r")
    # for line in inf.readlines():
    #     part_len_d[line.strip().split()[0]] = int(line.strip().split()[1])
    # inf.close()
    # if chro!="chrUn":
        
    #     ## didnt check the region whether cross the part chr ##
    #     ## this can be a bug ##
    #     ## if u find that this isnt right, u should fix ths bug ##
    #     if need_region[0]  > part_len_d[chro+"_part1"]:
    #         need_chr = chro+"_part2"
    #         need_region[0] = need_region[0]-part_len_d[chro+"_part1"]
    #         need_region[1] = need_region[1]-part_len_d[chro+"_part1"]
    #     else:
    #         need_chr = chro+"_part1"
    # else:
    #     pass
    

    ##check vcf.gz file
    if sv_file.endswith(".gz"):
        if os.path.exists(sv_file+".csi"):
            pass
        else:
            sys.stderr.write("didnt find index file for the vcf.gz file\n")
            exit(-1)
    else:
        sys.stderr.write("vcf file must be bgziped and indexed using bcftools\n")
        exit(-1)

    ##get the snp in the region
    
    child = subprocess.Popen("%s view -r %s:%d-%d %s"%(bcftools,need_chr,need_region[0],need_region[1],sv_file),shell=True,stdout=subprocess.PIPE)
    #########################SV template
    SV=[]
    SV_item={
        "type":"",
        "position":0,
        "length":0,
        "stat_in_each_sample":[
            
        ]
    }

    SV_INF=[]
    SV_INF_item={
        "chr":"",
        "position":0,
        "length":0,
        "ref":"",
        "alt":"",
        "ann":[]
        ##ann1: allele,Annotation , Gene_Name ,Feature_ID ,HGVS.c ,HGVS.p
    }
    ###########################SV TEMPLATE
    for line in child.stdout.readlines():
        line = line.decode("utf-8")
        if line.startswith("##"):
            pass
        elif line.startswith("#"):
            sv_sample_order=line.strip().split()[9:]
        else:
            
            # print(line)
            ls = line.strip().split()
            if int(ls[1]) >= need_region[0] and int(ls[1]) < need_region[1]:
                pass
            else:
                # print(need_region[0], need_region[1],ls[1])
                continue
            tmp_sv_item=copy.deepcopy(SV_item)
            tmp_sv_inf_item=copy.deepcopy(SV_INF_item)
            tmp_sv_item["position"] = int(ls[1]) #if ls[0].endswith("_part1") or ls[0].endswith("Un") else int(ls[1])+part_len_d[chro+"_part1"]
            # print(int(ls[1]), tmp_sv_item["position"])
            tmp_sv_inf_item["chr"]=chro
            tmp_sv_inf_item["position"] = int(ls[1]) #if ls[0].endswith("_part1") or ls[0].endswith("Un") else int(ls[1])+part_len_d[chro+"_part1"]
            tmp_sv_inf_item["ref"] = ls[3]
            tmp_sv_inf_item["alt"] = ls[4]
            pattern = re.compile(r".*SVTYPE=(.{,10});.*END=(\d+);.*") # INV INS DEL DUP
            match=pattern.match(ls[7])
            if not match:
                # print(ls[7])
                continue
            svtype=match.groups()[0]
            ed=int(match.groups()[1])

            tmp_sv_inf_item["type"] = svtype
            tmp_sv_item["type"] = svtype
            
            tmp_sv_inf_item["length"] = ed - int(ls[1])
            tmp_sv_item["length"] = ed - int(ls[1])
            
            
            for sample in ls[9:]:
                gt = sample.split(":")[0]
                if gt[0] == gt[-1]:
                    if gt[0] == "1":
                        tmp_sv_item["stat_in_each_sample"].append("1")
                    elif gt[0] == "0":
                        tmp_sv_item["stat_in_each_sample"].append("0")
                    else:
                        tmp_sv_item["stat_in_each_sample"].append("./.")
                else:
                    tmp_sv_item["stat_in_each_sample"].append("0/1")

            SV_INF.append(tmp_sv_inf_item)
            SV.append(tmp_sv_item)
        
    return (SV,SV_INF,sv_sample_order)



if __name__ == "__main__":
    cmdparser = argparse.ArgumentParser(description="SV and RDV to excel/fengcong@caas.cn" )
    #input data
    cmdparser.add_argument("-r","--region", dest="region",type=str, required=True,
                            help="plot region")
    cmdparser.add_argument("-g","--gff", dest="gff",type=str, required=False,
                            help="gff file if u want to plot gene structure")
    cmdparser.add_argument("-s","--sv", dest="sv",type=str, required=False,
                            help="sv file")
    ## No matter what, deep information must be drawn.
    cmdparser.add_argument("-d","--order", dest="order",type=str, required=True,
                            help="sample  order information ")
    #excel setting
    cmdparser.add_argument("-e","--excel", dest="excel",type=str, required=False,
                            help="input excel file.[optional]")
    cmdparser.add_argument("-c","--cellbp", dest="cellbp",type=int,default=1, required=False,
                                help="How many bp does a cell represent.default=1 [optional]")
    cmdparser.add_argument("-s","--startrowcol", dest="startrowcol",type=str, default="1,1" , required=False,
                            help="start row and col.default=1,1 [optional]")


    args = cmdparser.parse_args()

    #deal args
    ## region
    region = args.region
    chromosome = region.split(":")[0]
    start = int(region.split(":")[1].split("-")[0])
    end = int(region.split(":")[1].split("-")[1])
    ## gff
    gff = args.gff
    ## sv and order
    sv = args.sv
    order = args.order

    if args.excel:
        wb = openpyxl.load_workbook(args.excel)
    else:
        wb = openpyxl.Workbook()
    ws = wb.active
    
    #read order file
    order_list = []
    with open(order) as f:
        for line in f:
            order_list.append(line.strip())
    

    #read sv file
    SV_and_INF=deal_sv(chromosome,start,end,sv,part_len)
    sv_sample_order=SV_and_INF[2]

    